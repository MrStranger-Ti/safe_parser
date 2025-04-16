import logging
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

from src.excel.column import Column
from src.utils.tqdm import CREATING_EXCEL_CONFIG
from src import config

log = logging.getLogger(__name__)


class ExcelManager:
    def __init__(
        self,
        rows: list[dict[str, str]],
        template_path: Path | str = config.TEMPLATE_PATH,
        sheet_name: str = config.SHEET_NAME,
    ) -> None:
        self.rows: list[list[Column]] = []
        self.template_path: Path | str = template_path
        self.sheet_name: str = sheet_name
        self._workbook: Workbook | None = None
        self._sheet: Worksheet | None = None

        for row in rows:
            prepared_row = []
            for name, value in row.items():
                if name in config.COLS:
                    prepared_row.append(Column(data=value, num=config.COLS[name]))

            self.rows.append(prepared_row)

    def __enter__(self) -> "ExcelManager":
        self._workbook = self._get_workbook()
        self._sheet = self._get_sheet()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self._save_workbook()
        self._workbook.close()

    def _get_workbook(self) -> Workbook:
        return load_workbook(self.template_path)

    def _get_sheet(self) -> Worksheet:
        try:
            sheet = self._workbook[self.sheet_name]
        except KeyError as exc:
            log.error(
                f"{str(exc)}. Sheet name {config.SHEET_NAME} not found. Set correct in config.ini."
            )
            raise exc

        return sheet

    def _save_workbook(self) -> None:
        log.info("Saving file")
        try:
            self._workbook.save(config.RESULT_FILE_DIR)
        except PermissionError as exc:
            log.error(f"{str(exc)}. Close result excel file and try again.")
            raise exc

        log.info("Excel file has been saved")

    def insert_rows(self) -> None:
        log.info("Starting to prepare excel file")

        sheet = self._get_sheet()
        for row_num, row in enumerate(
            tqdm(self.rows, **CREATING_EXCEL_CONFIG), start=config.START_ROW
        ):
            for col in row:
                sheet.cell(row=row_num, column=col.num, value=col.data)
